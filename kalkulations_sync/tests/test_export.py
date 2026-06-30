import base64
import io
import json

import openpyxl

from odoo.exceptions import UserError, ValidationError

import re
import zipfile

from .common import KalkSyncBaseCase, _make_template_xlsx, _make_template_marker_xlsx
from ..models.sale_order import adjust_formula, sanitize_export_xlsx


class TestAdjustFormula(KalkSyncBaseCase):
    """Unit tests for the formula row-offset helper."""

    def test_basic_relative(self):
        self.assertEqual(adjust_formula('=B2*C2', 1), '=B3*C3')

    def test_multi_row_offset(self):
        self.assertEqual(adjust_formula('=B2+D5', 3), '=B5+D8')

    def test_absolute_row_unchanged(self):
        self.assertEqual(adjust_formula('=$B$2*C2', 1), '=$B$2*C3')

    def test_absolute_col_relative_row(self):
        self.assertEqual(adjust_formula('=$B2', 2), '=$B4')

    def test_non_formula_unchanged(self):
        self.assertEqual(adjust_formula('hello', 5), 'hello')
        self.assertIsNone(adjust_formula(None, 1))
        self.assertEqual(adjust_formula(42, 1), 42)

    def test_zero_offset_unchanged(self):
        formula = '=A1+B2'
        self.assertEqual(adjust_formula(formula, 0), formula)

    def test_sum_function(self):
        self.assertEqual(adjust_formula('=SUM(B2:B10)', 1), '=SUM(B3:B11)')


class TestFormulaCacheCleanup(KalkSyncBaseCase):
    """openpyxl leaves empty <v></v> nodes on formula cells, which corrupt the
    file for Excel and make our re-import read formula cells as None. The export
    must strip them and flag the workbook for full recalculation on open."""

    @staticmethod
    def _sheet_xml(b64):
        raw = base64.b64decode(b64)
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            return {
                'sheet': z.read('xl/worksheets/sheet1.xml').decode('utf-8'),
                'workbook': z.read('xl/workbook.xml').decode('utf-8'),
            }

    def test_no_empty_value_nodes_in_export(self):
        self.order.company_id.amf_kalksync_template = _make_template_xlsx(extra_formula=True)
        xml = self._sheet_xml(self.order._generate_kalkulation_excel())
        self.assertEqual(xml['sheet'].count('<v></v>'), 0)
        self.assertEqual(xml['sheet'].count('<v/>'), 0)
        # The formula itself must survive the cleanup.
        self.assertIn('<f>', xml['sheet'])

    def test_full_calc_on_load_set(self):
        self.order.company_id.amf_kalksync_template = _make_template_xlsx(extra_formula=True)
        xml = self._sheet_xml(self.order._generate_kalkulation_excel())
        self.assertIn('fullCalcOnLoad', xml['workbook'])

    def test_export_has_no_sharepoint_custom_part(self):
        # A template that carries SharePoint custom properties must export
        # without docProps/custom.xml (Excel rejects openpyxl's rewrite of it).
        b64 = _make_template_xlsx(extra_formula=True)
        raw = base64.b64decode(b64)
        buf = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw)) as zin, \
                zipfile.ZipFile(buf, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == '[Content_Types].xml':
                    data = data.replace(
                        b'</Types>',
                        b'<Override PartName="/docProps/custom.xml" '
                        b'ContentType="application/vnd.openxmlformats-'
                        b'officedocument.custom-properties+xml"/></Types>',
                    )
                zout.writestr(item, data)
            zout.writestr(
                'docProps/custom.xml',
                '<Properties xmlns="http://schemas.openxmlformats.org/'
                'officeDocument/2006/custom-properties"><property name='
                '"ContentTypeId" fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}"'
                ' pid="2"><vt:lpwstr xmlns:vt="http://schemas.openxmlformats.org'
                '/officeDocument/2006/docPropsVTypes">0x0101</vt:lpwstr>'
                '</property></Properties>',
            )
        self.order.company_id.amf_kalksync_template = base64.b64encode(buf.getvalue())
        out = base64.b64decode(self.order._generate_kalkulation_excel())
        with zipfile.ZipFile(io.BytesIO(out)) as z:
            self.assertNotIn('docProps/custom.xml', z.namelist())
            self.assertNotIn('docProps/custom.xml', z.read('[Content_Types].xml').decode())

    def test_sanitize_keeps_real_caches_and_drops_sharepoint(self):
        # A real cached value must be preserved; only empty nodes are removed.
        # The SharePoint custom-properties part and its references are dropped.
        sheet = (
            '<worksheet><sheetData>'
            '<c r="A1"><f>SUM(B1:B2)</f><v></v></c>'
            '<c r="A2"><f>1+1</f><v>2</v></c>'
            '<c r="A3"><f>C1</f><v/></c>'
            '</sheetData></worksheet>'
        )
        content_types = (
            '<Types><Override PartName="/docProps/custom.xml" ContentType="x"/>'
            '<Override PartName="/xl/styles.xml" ContentType="y"/></Types>'
        )
        rels = (
            '<Relationships>'
            '<Relationship Id="rId4" Type="t" Target="docProps/custom.xml"/>'
            '<Relationship Id="rId1" Type="t" Target="xl/workbook.xml"/>'
            '</Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w') as z:
            z.writestr('xl/worksheets/sheet1.xml', sheet)
            z.writestr('docProps/custom.xml', '<Properties/>')
            z.writestr('[Content_Types].xml', content_types)
            z.writestr('_rels/.rels', rels)
        core = (
            '<cp:coreProperties><dcterms:modified>'
            '2026-06-18T10:00:00+00:00Z</dcterms:modified>'
            '<dcterms:created>2000-11-26T14:42:47Z</dcterms:created>'
            '</cp:coreProperties>'
        )
        with zipfile.ZipFile(buf, 'a') as z:
            z.writestr('docProps/core.xml', core)
        out = sanitize_export_xlsx(buf.getvalue())
        with zipfile.ZipFile(io.BytesIO(out)) as z:
            names = z.namelist()
            res = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
            ct = z.read('[Content_Types].xml').decode('utf-8')
            rl = z.read('_rels/.rels').decode('utf-8')
            co = z.read('docProps/core.xml').decode('utf-8')
        # invalid "+00:00Z" repaired to valid "+00:00"; valid "…Z" left alone
        self.assertNotIn('+00:00Z', co)
        self.assertIn('2026-06-18T10:00:00+00:00<', co)
        self.assertIn('2000-11-26T14:42:47Z', co)
        # empty caches gone, real cache + all formulas kept
        self.assertEqual(res.count('<v></v>'), 0)
        self.assertEqual(res.count('<v/>'), 0)
        self.assertIn('<v>2</v>', res)
        self.assertEqual(res.count('<f>'), 3)
        # SharePoint part and its references removed, others untouched
        self.assertNotIn('docProps/custom.xml', names)
        self.assertNotIn('docProps/custom.xml', ct)
        self.assertNotIn('docProps/custom.xml', rl)
        self.assertIn('/xl/styles.xml', ct)
        self.assertIn('xl/workbook.xml', rl)


class TestExportLineCount(KalkSyncBaseCase):
    """Test that the exported file contains the right number of data rows."""

    def _read_exported_wb(self):
        b64 = self.order._generate_kalkulation_excel()
        return openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(b64)), data_only=False
        )

    def test_line_count(self):
        wb = self._read_exported_wb()
        ws = wb.active
        # Template has header rows 1+2, data starts at row 3
        # We have 2 SO lines → rows 3 and 4
        data_rows = [
            row for row in ws.iter_rows(min_row=3, values_only=True)
            if any(v is not None for v in row)
        ]
        self.assertEqual(len(data_rows), 2)

    def test_ids_written(self):
        wb = self._read_exported_wb()
        ws = wb.active
        id_values = [ws.cell(row=3 + i, column=1).value for i in range(2)]
        expected = sorted([self.line1.id, self.line2.id])
        self.assertEqual(sorted(id_values), expected)

    def test_no_placeholders_remain(self):
        wb = self._read_exported_wb()
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    self.assertNotIn('{{', cell, "Placeholder not replaced")

    def test_formula_offset_applied(self):
        self.order.company_id.amf_kalksync_template = _make_template_xlsx(extra_formula=True)
        wb = self._read_exported_wb()
        ws = wb.active
        # Row 3: formula should reference row 3; row 4: row 4
        formula_row3 = ws.cell(row=3, column=5).value
        formula_row4 = ws.cell(row=4, column=5).value
        self.assertIn('C3', str(formula_row3))
        self.assertIn('D3', str(formula_row3))
        self.assertIn('C4', str(formula_row4))
        self.assertIn('D4', str(formula_row4))

    def test_export_blocked_on_sale_state(self):
        self.order.action_confirm()
        with self.assertRaises(UserError):
            self.order.action_export_kalkulation()

    def test_export_blocked_without_template(self):
        self.order.company_id.amf_kalksync_template = False
        with self.assertRaises(UserError):
            self.order.action_export_kalkulation()
        # restore
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_export_blocked_without_line_id_placeholder(self):
        self.order.company_id.amf_kalksync_template = _make_template_xlsx(include_line_id=False)
        with self.assertRaises(ValidationError):
            self.order.action_export_kalkulation()
        # restore
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_meta_sheet_present(self):
        wb = self._read_exported_wb()
        self.assertIn('kalksync_meta', wb.sheetnames)

    def test_meta_sale_order_id(self):
        wb = self._read_exported_wb()
        meta = wb['kalksync_meta']
        self.assertEqual(int(meta['B2'].value), self.order.id)

    def test_marker_only_column_in_mapping(self):
        """Variante A: [marker] in header row registers field without a placeholder."""
        self.order.company_id.amf_kalksync_template = _make_template_marker_xlsx()
        b64 = self.order._generate_kalkulation_excel()
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        col_mapping = json.loads(wb['kalksync_meta']['B5'].value)
        self.assertIn('name', col_mapping, "Marker-only column must appear in col_mapping")
        # restore
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_header_object_placeholder_replaced(self):
        """{{object.partner_id.name}} in header rows is replaced with the partner's name."""
        wb = self._read_exported_wb()
        ws = wb.active
        # D1 has {{object.partner_id.name}} → should equal 'Testkunde'
        self.assertEqual(ws['D1'].value, self.partner.name)

    def test_export_single_line_no_insert_rows(self):
        """Export with exactly 1 order line uses the master row without insert_rows."""
        so = self.env['sale.order'].create({'partner_id': self.partner.id})
        so.company_id.amf_kalksync_template = _make_template_xlsx()
        self.env['sale.order.line'].create({
            'order_id': so.id,
            'product_id': self.product.id,
            'product_uom_qty': 3.0,
            'price_unit': 15.0,
        })
        b64 = so._generate_kalkulation_excel()
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        ws = wb.active
        data_rows = [
            row for row in ws.iter_rows(min_row=3, values_only=True)
            if any(v is not None for v in row)
        ]
        self.assertEqual(len(data_rows), 1)

    def test_export_no_importable_lines_raises(self):
        """Export with no importable order lines (only sections/notes) raises UserError."""
        so = self.env['sale.order'].create({'partner_id': self.partner.id})
        so.company_id.amf_kalksync_template = _make_template_xlsx()
        with self.assertRaises(UserError):
            so._generate_kalkulation_excel()
