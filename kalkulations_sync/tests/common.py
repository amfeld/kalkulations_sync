import base64
import io

import openpyxl

from odoo.tests.common import TransactionCase


def _make_template_xlsx(include_line_id=True, extra_formula=False):
    """Build a minimal .xlsx template.

    {{line.field}} placeholders in the master row are export-only (display).
    Import columns are declared via [marker] headers in row 2 — here for
    qty, price and name, so the standard import tests can patch those columns.
    Returns base64-encoded bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kalkulation'

    # Header row
    ws['A1'] = 'Angebot:'
    ws['B1'] = '{{object.name}}'
    ws['C1'] = 'Kunde:'
    ws['D1'] = '{{object.partner_id.name}}'

    # Column headers (row 2) — [marker] declares the column as import source
    ws['A2'] = 'ID'
    ws['B2'] = 'Produkt'
    ws['C2'] = 'Menge [product_uom_qty]'
    ws['D2'] = 'VK-Preis [price_unit]'
    ws['E2'] = 'Summe'
    ws['F2'] = 'Bezeichnung [name]'

    # Master row (row 3) — placeholders fill values on export (display only)
    if include_line_id:
        ws['A3'] = '{{line.id}}'
    ws['B3'] = '{{line.product_id.name}}'
    ws['C3'] = '{{line.product_uom_qty}}'
    ws['D3'] = '{{line.price_unit}}'
    ws['F3'] = '{{line.name}}'
    if extra_formula:
        ws['E3'] = '=C3*D3'

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def _make_template_marker_xlsx():
    """Build a template using the [marker] mechanism (Variante A) for column F.

    Column F has a [name] header marker in row 2 but no placeholder in the
    master row — exercising the header-scan code path independently.
    Returns base64-encoded bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kalkulation'

    ws['A1'] = 'Angebot:'
    ws['B1'] = '{{object.name}}'

    # Marker row (row 2): [name] registers column F as importable via Variante A
    ws['A2'] = 'ID'
    ws['B2'] = 'Produkt'
    ws['C2'] = 'Menge'
    ws['D2'] = 'VK-Preis'
    ws['F2'] = '[name]'

    # Master row (row 3): column F has no placeholder — value stays blank on export
    ws['A3'] = '{{line.id}}'
    ws['B3'] = '{{line.product_id.name}}'
    ws['C3'] = '{{line.product_uom_qty}}'
    ws['D3'] = '{{line.price_unit}}'

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


class KalkSyncBaseCase(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # Product
        cls.product = cls.env['product.product'].create({
            'name': 'Testprodukt',
            'type': 'consu',
            'list_price': 100.0,
        })

        # Partner
        cls.partner = cls.env['res.partner'].create({'name': 'Testkunde'})

        # Sale order
        cls.order = cls.env['sale.order'].create({
            'partner_id': cls.partner.id,
        })
        cls.line1 = cls.env['sale.order.line'].create({
            'order_id': cls.order.id,
            'product_id': cls.product.id,
            'product_uom_qty': 2.0,
            'price_unit': 50.0,
        })
        cls.line2 = cls.env['sale.order.line'].create({
            'order_id': cls.order.id,
            'product_id': cls.product.id,
            'product_uom_qty': 5.0,
            'price_unit': 20.0,
        })

        # Mark both lines as carrying a *manual* price. This module exists to protect
        # manually calculated prices; a real Kalkulation line always has one. In Odoo
        # v19 a create({'price_unit': X}) precomputes technical_price_unit == X, so
        # has_manual_price() reads False and a later qty write would silently reset
        # price_unit back to list_price. Pushing technical_price_unit to the automatic
        # value (list_price) makes the guard recognise the manual override, matching
        # the real-world fixture and keeping unrelated fields stable on import.
        cls.env.cr.execute(
            "UPDATE sale_order_line SET technical_price_unit = %s WHERE id IN %s",
            [cls.product.list_price, tuple((cls.line1 | cls.line2).ids)],
        )
        cls.env.invalidate_all()

        # Store template on company
        cls.order.company_id.amf_kalksync_template = _make_template_xlsx()
        cls.order.company_id.amf_kalksync_template_name = 'test_template.xlsx'
        # Default product used when 'N' rows create new lines without a product
        cls.order.company_id.amf_kalksync_default_product_id = cls.product.id
