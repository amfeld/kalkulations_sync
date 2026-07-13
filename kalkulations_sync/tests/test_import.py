import base64
import io
import json

import openpyxl
from odoo import fields
from odoo.exceptions import UserError

from .common import KalkSyncBaseCase


def _patch_exported_file(b64_file, line_id, col_letter, new_value):
    """Helper: open an exported xlsx, change one cell, return new base64."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    ws = wb.active
    meta = wb['kalksync_meta']
    data_row_start = int(meta['B6'].value)
    id_col = meta['B4'].value

    from openpyxl.utils import column_index_from_string
    id_col_idx = column_index_from_string(id_col)

    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            from openpyxl.utils import column_index_from_string as c2i
            row[c2i(col_letter) - 1].value = new_value
            break

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def _corrupt_id(b64_file, line_id, new_id_value):
    """Replace a line ID with a corrupt value."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    ws = wb.active
    meta = wb['kalksync_meta']
    data_row_start = int(meta['B6'].value)
    id_col = meta['B4'].value

    from openpyxl.utils import column_index_from_string
    id_col_idx = column_index_from_string(id_col)

    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            row[id_col_idx - 1].value = new_id_value
            break

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def _clear_line_row(b64_file, line_id):
    """Blank out all cells in the row that belongs to line_id, simulating a missing line."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    ws = wb.active
    meta = wb['kalksync_meta']
    data_row_start = int(meta['B6'].value)
    from openpyxl.utils import column_index_from_string
    id_col_idx = column_index_from_string(meta['B4'].value)
    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            for cell in row:
                cell.value = None
            break
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


class TestImportIdMismatch(KalkSyncBaseCase):

    def _make_wizard(self, file_b64, file_name='test.xlsx'):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': file_name,
        })
        # Trigger onchange manually in tests
        wizard._onchange_file_data()
        return wizard

    def test_corrupted_id_gets_error_status(self):
        b64 = self.order._generate_kalkulation_excel()
        corrupted = _corrupt_id(b64, self.line1.id, 'MANIPULIERT')
        wizard = self._make_wizard(corrupted)
        error_lines = wizard.line_ids.filtered(lambda l: l.status == 'error')
        self.assertTrue(error_lines, "Manipulierte ID muss error-Status erhalten")

    def test_has_errors_blocks_confirm(self):
        b64 = self.order._generate_kalkulation_excel()
        corrupted = _corrupt_id(b64, self.line1.id, 'BAD')
        wizard = self._make_wizard(corrupted)
        self.assertTrue(wizard.has_errors)
        with self.assertRaises(UserError):
            wizard.action_confirm()

    def test_unknown_id_gets_error_status(self):
        b64 = self.order._generate_kalkulation_excel()
        corrupted = _corrupt_id(b64, self.line1.id, 999999)
        wizard = self._make_wizard(corrupted)
        error_lines = wizard.line_ids.filtered(lambda l: l.status == 'error')
        self.assertTrue(error_lines)

    def test_missing_meta_sheet_raises(self):
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'no meta'
        buf = io.BytesIO()
        wb.save(buf)
        bad_b64 = base64.b64encode(buf.getvalue()).decode()
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': bad_b64,
            'file_name': 'bad.xlsx',
        })
        result = wizard._onchange_file_data()
        self.assertIn('warning', result or {})

    def test_wrong_sale_order_raises(self):
        other_order = self.env['sale.order'].create({'partner_id': self.partner.id})
        self.env['sale.order.line'].create({
            'order_id': other_order.id,
            'product_id': self.product.id,
            'product_uom_qty': 1.0,
            'price_unit': 10.0,
        })
        other_order.company_id.amf_kalksync_template = self.order.company_id.amf_kalksync_template
        b64 = other_order._generate_kalkulation_excel()
        # Try to import the other order's file into self.order
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64,
            'file_name': 'other.xlsx',
        })
        result = wizard._onchange_file_data()
        self.assertIn('warning', result or {})


class TestNewLine(KalkSyncBaseCase):

    def _make_wizard(self, file_b64, file_name='test.xlsx'):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': file_name,
        })
        wizard._onchange_file_data()
        return wizard

    def test_n_row_gets_new_status(self):
        b64 = self.order._generate_kalkulation_excel()
        b64_n = _corrupt_id(b64, self.line1.id, 'N')
        wizard = self._make_wizard(b64_n)
        new_lines = wizard.line_ids.filtered(lambda l: l.status == 'new')
        self.assertTrue(new_lines, "N in ID-Spalte muss als 'new' erkannt werden")

    def test_n_lowercase_gets_new_status(self):
        b64 = self.order._generate_kalkulation_excel()
        b64_n = _corrupt_id(b64, self.line1.id, 'n')
        wizard = self._make_wizard(b64_n)
        new_lines = wizard.line_ids.filtered(lambda l: l.status == 'new')
        self.assertTrue(new_lines, "Kleinbuchstabe n muss ebenfalls erkannt werden")

    def test_n_row_does_not_block_confirm(self):
        b64 = self.order._generate_kalkulation_excel()
        b64_n = _corrupt_id(b64, self.line1.id, 'N')
        wizard = self._make_wizard(b64_n)
        self.assertFalse(wizard.has_errors)

    def test_n_row_creates_new_line(self):
        b64 = self.order._generate_kalkulation_excel()
        b64_n = _corrupt_id(b64, self.line1.id, 'N')
        count_before = len(self.order.order_line)
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64_n,
            'file_name': 'new_line.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()
        self.assertEqual(len(self.order.order_line), count_before + 1)

    def test_n_row_name_from_column(self):
        b64 = self.order._generate_kalkulation_excel()
        # Corrupt line1 ID to 'N'; exported F-column already has line1.name
        b64_n = _corrupt_id(b64, self.line1.id, 'N')
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64_n,
            'file_name': 'new_named.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()
        new_line = self.order.order_line.filtered(
            lambda l: l.id not in [self.line1.id, self.line2.id]
        )
        self.assertEqual(len(new_line), 1)
        self.assertEqual(new_line.name, self.line1.name)


class TestImportRoundtrip(KalkSyncBaseCase):

    def test_changed_price_unit_imported(self):
        b64 = self.order._generate_kalkulation_excel()

        # Find which column is price_unit
        wb_check = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        col_mapping = json.loads(wb_check['kalksync_meta']['B5'].value)
        price_col = col_mapping.get('price_unit')
        self.assertTrue(price_col, "price_unit must be in column_mapping")

        patched = _patch_exported_file(b64, self.line1.id, price_col, 99.0)

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': patched,
            'file_name': 'patched.xlsx',
        })
        wizard._onchange_file_data()

        changed = wizard.line_ids.filtered(
            lambda l: l.status == 'changed' and l.field_name == 'price_unit'
                      and l.order_line_id == self.line1
        )
        self.assertTrue(changed)
        self.assertEqual(changed[0].value_new, '99')

        wizard.action_confirm()
        self.assertAlmostEqual(self.line1.price_unit, 99.0)

    def test_unchanged_line_not_written(self):
        original_price = self.line1.price_unit
        b64 = self.order._generate_kalkulation_excel()

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64,
            'file_name': 'noop.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()
        self.assertAlmostEqual(self.line1.price_unit, original_price)

    def test_chatter_message_posted(self):
        b64 = self.order._generate_kalkulation_excel()
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64,
            'file_name': 'test.xlsx',
        })
        wizard._onchange_file_data()
        msg_count_before = len(self.order.message_ids)
        wizard.action_confirm()
        self.assertGreater(len(self.order.message_ids), msg_count_before)
        # Body must arrive as real HTML, not double-escaped text.
        body = self.order.message_ids[0].body
        self.assertIn('<p>', body)
        self.assertNotIn('&lt;', body)

    def test_chatter_message_groups_changes_per_line(self):
        b64 = self.order._generate_kalkulation_excel()
        wb_check = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        col_mapping = json.loads(wb_check['kalksync_meta']['B5'].value)
        patched = _patch_exported_file(b64, self.line1.id, col_mapping['price_unit'], 99.0)

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': patched,
            'file_name': 'patched.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()

        body = self.order.message_ids[0].body
        # Line heading (first description line, bold) + bullet list of changes
        self.assertIn('<strong>%s</strong>' % self.line1.name.splitlines()[0], body)
        self.assertIn('<li>', body)
        self.assertNotIn('&lt;', body)


class TestMissingLine(KalkSyncBaseCase):

    def _make_wizard(self, file_b64, file_name='test.xlsx'):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': file_name,
        })
        wizard._onchange_file_data()
        return wizard

    def test_missing_line_gets_missing_status(self):
        """A non-section SO line absent from Excel must appear with 'missing' status."""
        b64 = self.order._generate_kalkulation_excel()
        b64_cleared = _clear_line_row(b64, self.line1.id)
        wizard = self._make_wizard(b64_cleared)
        missing = wizard.line_ids.filtered(lambda l: l.status == 'missing')
        self.assertTrue(missing, "Lines absent from Excel must get 'missing' status")

    def test_duplicate_id_gets_new_status(self):
        """A line_id repeated twice in Excel (user-copied row) is treated as a new line."""
        b64 = self.order._generate_kalkulation_excel()
        # Make line2 carry line1's ID → duplicate
        b64_dup = _corrupt_id(b64, self.line2.id, self.line1.id)
        wizard = self._make_wizard(b64_dup)
        new_lines = wizard.line_ids.filtered(lambda l: l.status == 'new')
        self.assertTrue(new_lines, "Duplicate ID must be treated as a new copied line")


class TestNewLineEdgeCases(KalkSyncBaseCase):

    def _make_wizard(self, file_b64, file_name='test.xlsx'):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': file_name,
        })
        wizard._onchange_file_data()
        return wizard

    def test_n_row_without_default_product_raises(self):
        """Confirming N-row without amf_kalksync_default_product_id configured raises UserError."""
        self.order.company_id.amf_kalksync_default_product_id = False
        b64 = self.order._generate_kalkulation_excel()
        b64_n = _corrupt_id(b64, self.line1.id, 'N')
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64_n,
            'file_name': 'new_line.xlsx',
        })
        wizard._onchange_file_data()
        with self.assertRaises(UserError):
            wizard.action_confirm()


class TestImportActions(KalkSyncBaseCase):

    def test_action_import_kalkulation_opens_wizard(self):
        """action_import_kalkulation returns an act_window pointing at the import wizard."""
        result = self.order.action_import_kalkulation()
        self.assertEqual(result['type'], 'ir.actions.act_window')
        self.assertEqual(result['res_model'], 'kalksync.import.wizard')
        self.assertEqual(result['context']['default_sale_order_id'], self.order.id)
