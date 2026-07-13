"""Selection-field validation on import + marker-only column mapping.

Covers the 2.x behaviour change:
- Import columns come exclusively from [marker] headers; {{line.field}}
  placeholders in the master row are export-only (display).
- Selection values from Excel are validated against the target instance's
  allowed keys (key or translated label accepted) instead of being written
  raw — an invalid value becomes a per-row error, not a ValueError that
  rolls back the whole confirm.
"""
import base64
import io
import json

import openpyxl
from openpyxl.utils import column_index_from_string

from .common import KalkSyncBaseCase, _make_template_xlsx


def _add_selection_column(b64_file, field_name, col_letter, line_id, cell_value):
    """Register field_name→col_letter in kalksync_meta and write cell_value
    into that column on the row belonging to line_id."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    ws = wb.active
    meta = wb['kalksync_meta']
    mapping = json.loads(meta['B5'].value or '{}')
    mapping[field_name] = col_letter
    meta['B5'] = json.dumps(mapping)

    data_row_start = int(meta['B6'].value)
    id_col_idx = column_index_from_string(meta['B4'].value)
    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            ws.cell(row=row[0].row, column=column_index_from_string(col_letter),
                    value=cell_value)
            break
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


class TestMarkerOnlyMapping(KalkSyncBaseCase):
    """{{line.field}} placeholders must no longer register import columns."""

    def test_placeholder_only_template_yields_empty_mapping(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '{{line.id}}'
        ws['B1'] = '{{line.product_uom_qty}}'
        ws['C1'] = '{{line.price_unit}}'
        buf = io.BytesIO()
        wb.save(buf)
        self.order.company_id.amf_kalksync_template = \
            base64.b64encode(buf.getvalue()).decode()

        b64 = self.order._generate_kalkulation_excel()
        out = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
        mapping = json.loads(out['kalksync_meta']['B5'].value or '{}')
        self.assertEqual(
            mapping, {},
            "Placeholders in the master row must not create import mappings",
        )
        # restore for sibling tests
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_marker_beats_placeholder_column(self):
        """Display column (placeholder) and upload column (marker) for the
        same field: the mapping must point at the marker column."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['D1'] = '[price_unit]'
        ws['A2'] = '{{line.id}}'
        ws['B2'] = '{{line.price_unit}}'   # display only
        ws['D2'] = '=B2*1.15'              # upload column under the marker
        buf = io.BytesIO()
        wb.save(buf)
        self.order.company_id.amf_kalksync_template = \
            base64.b64encode(buf.getvalue()).decode()

        b64 = self.order._generate_kalkulation_excel()
        out = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
        mapping = json.loads(out['kalksync_meta']['B5'].value or '{}')
        self.assertEqual(mapping.get('price_unit'), 'D')
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_export_warns_when_no_markers(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '{{line.id}}'
        ws['B1'] = '{{line.price_unit}}'
        buf = io.BytesIO()
        wb.save(buf)
        self.order.company_id.amf_kalksync_template = \
            base64.b64encode(buf.getvalue()).decode()
        self.order.company_id.amf_kalksync_template_name = 'no_markers.xlsx'

        before = self.order.message_ids
        self.order.action_export_kalkulation()
        new_msgs = self.order.message_ids - before
        self.assertTrue(new_msgs)
        self.assertIn('import markers', new_msgs[0].body)
        # The warning must survive message_post's HTML handling as real
        # markup (own line), not as escaped text glued to the export note.
        self.assertIn('<br', str(new_msgs[0].body))
        self.assertNotIn('&lt;br', str(new_msgs[0].body))
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()

    def test_marker_and_placeholder_same_column(self):
        """Real-world starter-template pattern: display placeholder AND
        upload marker share one column — mapping points at that column."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['D1'] = '[price_unit]'
        ws['A2'] = '{{line.id}}'
        ws['D2'] = '{{line.price_unit}}'
        buf = io.BytesIO()
        wb.save(buf)
        self.order.company_id.amf_kalksync_template = \
            base64.b64encode(buf.getvalue()).decode()

        b64 = self.order._generate_kalkulation_excel()
        out = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
        mapping = json.loads(out['kalksync_meta']['B5'].value or '{}')
        self.assertEqual(mapping.get('price_unit'), 'D')
        # Export still fills the placeholder value in the marked column
        ws_out = out.active
        self.assertEqual(ws_out['D2'].value, self.line1.price_unit)
        self.order.company_id.amf_kalksync_template = _make_template_xlsx()


class TestSelectionValidation(KalkSyncBaseCase):

    def _make_wizard(self, file_b64):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': 'test.xlsx',
        })
        wizard._onchange_file_data()
        return wizard

    def test_selection_key_accepts_key_label_and_rejects_junk(self):
        wiz = self.env['kalksync.import.wizard'].new({
            'sale_order_id': self.order.id,
        })
        f = self.env['sale.order.line']._fields['display_type']
        pairs = f._description_selection(self.env)
        key, label = pairs[0]
        self.assertEqual(wiz._selection_key(f, key), key)
        self.assertEqual(wiz._selection_key(f, str(label)), key)
        self.assertEqual(wiz._selection_key(f, f'  {str(label).upper()}  '), key)
        self.assertIsNone(wiz._selection_key(f, 'definitely_not_a_key'))

    def test_invalid_selection_value_becomes_error_row(self):
        b64 = self.order._generate_kalkulation_excel()
        patched = _add_selection_column(
            b64, 'display_type', 'H', self.line1.id, 'normal',
        )
        wizard = self._make_wizard(patched)
        err = wizard.line_ids.filtered(
            lambda l: l.field_name == 'display_type' and l.status == 'error'
        )
        self.assertTrue(
            err, "Invalid selection value must produce an error row, not crash",
        )
        self.assertIn('normal', err[0].error_message)

    def test_valid_selection_key_maps_to_change(self):
        b64 = self.order._generate_kalkulation_excel()
        line_vals_key = 'line_note'
        patched = _add_selection_column(
            b64, 'display_type', 'H', self.line1.id, line_vals_key,
        )
        wizard = self._make_wizard(patched)
        chg = wizard.line_ids.filtered(
            lambda l: l.field_name == 'display_type' and l.status == 'changed'
        )
        self.assertTrue(chg)
        # Parse returns the mapped key in _raw_value for confirm
        line_vals, _dt, _w = wizard._parse_excel()
        raw = [
            v for v in line_vals
            if v.get('field_name') == 'display_type'
            and v.get('status') == 'changed'
        ]
        self.assertTrue(raw)
        self.assertEqual(raw[0]['_raw_value'], line_vals_key)

    def test_blank_selection_cell_is_no_change(self):
        b64 = self.order._generate_kalkulation_excel()
        patched = _add_selection_column(
            b64, 'display_type', 'H', self.line1.id, None,
        )
        wizard = self._make_wizard(patched)
        rows = wizard.line_ids.filtered(
            lambda l: l.field_name == 'display_type'
        )
        self.assertFalse(
            rows, "Blank selection cell must be skipped entirely",
        )
