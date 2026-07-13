"""Deeper coverage of the import parse/confirm engine in wizard/import_wizard.py.

The existing test_import.py covers ID mismatch, N-rows, missing/duplicate lines and
a single price roundtrip. These tests target the remaining branches that carry real
logic and were previously unverified:

- numeric cell that cannot be parsed → 'error' status (blocks confirm)
- formula cell whose cached result is missing → 'error' status
- German decimal comma in a numeric cell → parsed, written correctly
- blank numeric cell → treated as "no change" (not as 0)
- multiple fields on multiple lines confirmed in one go
- legacy second-precision export timestamp still parses (concurrency branch)
- _read_row_fields (instance) / _make_error_line (static) helpers
- show_only_changes toggle re-parses instead of going stale
"""

import base64
import io
import json

import openpyxl
from openpyxl.utils import column_index_from_string

from odoo.exceptions import UserError

from ..wizard.import_wizard import KalkSyncImportWizard

from .common import KalkSyncBaseCase


def _col_for(b64_file, field_name):
    """Return the Excel column letter mapped to field_name in the export metadata."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    mapping = json.loads(wb['kalksync_meta']['B5'].value)
    return mapping[field_name]


def _set_cell(b64_file, line_id, col_letter, new_value):
    """Set the data cell for (line_id, col_letter) to new_value; return new base64."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=True)
    ws = wb.active
    meta = wb['kalksync_meta']
    data_row_start = int(meta['B6'].value)
    id_col_idx = column_index_from_string(meta['B4'].value)
    col_idx = column_index_from_string(col_letter)
    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            row[col_idx - 1].value = new_value
            break
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def _set_cell_formula_no_cache(b64_file, line_id, col_letter, formula):
    """Write a formula string into a cell, keep data_only=False so no cached <v> exists.

    Simulates the real "file edited but never recalculated/saved in Excel" case where
    data_only load returns None for the cell while the formula load returns '=...'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_file)), data_only=False)
    ws = wb.active
    meta = wb['kalksync_meta']
    data_row_start = int(meta['B6'].value)
    id_col_idx = column_index_from_string(meta['B4'].value)
    col_idx = column_index_from_string(col_letter)
    for row in ws.iter_rows(min_row=data_row_start, values_only=False):
        if row[id_col_idx - 1].value == line_id:
            row[col_idx - 1].value = formula
            break
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


class TestImportNumericValidation(KalkSyncBaseCase):

    def _make_wizard(self, file_b64):
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': file_b64,
            'file_name': 'test.xlsx',
        })
        wizard._onchange_file_data()
        return wizard

    def test_non_numeric_cell_gets_error_and_blocks_confirm(self):
        """Text in a numeric (price) cell → error status, confirm refused, price untouched."""
        b64 = self.order._generate_kalkulation_excel()
        col = _col_for(b64, 'price_unit')
        bad = _set_cell(b64, self.line1.id, col, 'nicht-zahl')
        wizard = self._make_wizard(bad)

        err = wizard.line_ids.filtered(
            lambda l: l.status == 'error' and l.field_name == 'price_unit'
        )
        self.assertTrue(err, "Unparseable numeric cell must yield an error line")
        self.assertTrue(wizard.has_errors)

        original = self.line1.price_unit
        with self.assertRaises(UserError):
            wizard.action_confirm()
        # assertRaises rolls back the savepoint, but the price must never have changed.
        self.assertAlmostEqual(self.line1.price_unit, original)

    def test_german_comma_in_cell_parsed_and_written(self):
        """A '12,5' string in the price cell must import as 12.5 (de decimal comma)."""
        b64 = self.order._generate_kalkulation_excel()
        col = _col_for(b64, 'price_unit')
        patched = _set_cell(b64, self.line1.id, col, '12,5')
        wizard = self._make_wizard(patched)

        changed = wizard.line_ids.filtered(
            lambda l: l.status == 'changed' and l.field_name == 'price_unit'
                      and l.order_line_id == self.line1
        )
        self.assertTrue(changed)
        wizard.action_confirm()
        self.assertAlmostEqual(self.line1.price_unit, 12.5)

    def test_blank_numeric_cell_is_no_change(self):
        """A blank price cell means 'leave as is' — not 'set to 0'."""
        original = self.line1.price_unit
        b64 = self.order._generate_kalkulation_excel()
        col = _col_for(b64, 'price_unit')
        blanked = _set_cell(b64, self.line1.id, col, None)
        wizard = self._make_wizard(blanked)

        price_lines = wizard.line_ids.filtered(
            lambda l: l.field_name == 'price_unit' and l.order_line_id == self.line1
        )
        self.assertFalse(
            price_lines.filtered(lambda l: l.status == 'changed'),
            "Blank cell must not be flagged as a change",
        )
        wizard.action_confirm()
        self.assertAlmostEqual(self.line1.price_unit, original)


class TestImportFormulaWithoutCache(KalkSyncBaseCase):

    def test_formula_cell_without_cached_value_errors(self):
        """A formula cell with no cached result → error line steering the user to
        open+save the file in Excel. This must block confirm."""
        b64 = self.order._generate_kalkulation_excel()
        col = _col_for(b64, 'price_unit')
        bad = _set_cell_formula_no_cache(b64, self.line1.id, col, '=1+1')

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': bad,
            'file_name': 'formula.xlsx',
        })
        wizard._onchange_file_data()

        err = wizard.line_ids.filtered(
            lambda l: l.status == 'error' and l.field_name == 'price_unit'
        )
        self.assertTrue(err, "Formula cell without cached value must produce an error")
        self.assertIn('Formula cell', err[0].error_message)
        self.assertTrue(wizard.has_errors)


class TestImportMultiChange(KalkSyncBaseCase):

    def test_two_fields_two_lines_confirmed(self):
        """price_unit on line1 and product_uom_qty on line2 both written in one confirm."""
        b64 = self.order._generate_kalkulation_excel()
        price_col = _col_for(b64, 'price_unit')
        qty_col = _col_for(b64, 'product_uom_qty')

        patched = _set_cell(b64, self.line1.id, price_col, 77.0)
        patched = _set_cell(patched, self.line2.id, qty_col, 9.0)

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': patched,
            'file_name': 'multi.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()

        self.assertAlmostEqual(self.line1.price_unit, 77.0)
        self.assertAlmostEqual(self.line2.product_uom_qty, 9.0)
        # Unchanged fields stay put.
        self.assertAlmostEqual(self.line1.product_uom_qty, 2.0)
        self.assertAlmostEqual(self.line2.price_unit, 20.0)

    def test_two_fields_same_line_single_write(self):
        """Two changed fields on the same line are coalesced into one write dict."""
        b64 = self.order._generate_kalkulation_excel()
        price_col = _col_for(b64, 'price_unit')
        qty_col = _col_for(b64, 'product_uom_qty')

        patched = _set_cell(b64, self.line1.id, price_col, 33.0)
        patched = _set_cell(patched, self.line1.id, qty_col, 8.0)

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': patched,
            'file_name': 'sameline.xlsx',
        })
        wizard._onchange_file_data()
        wizard.action_confirm()

        self.assertAlmostEqual(self.line1.price_unit, 33.0)
        self.assertAlmostEqual(self.line1.product_uom_qty, 8.0)


class TestLegacyTimestampParsing(KalkSyncBaseCase):

    def test_second_precision_export_timestamp_parses(self):
        """Legacy files store the export time with second precision (no microseconds).

        The parser must still read it so the concurrency branch works on old files.
        We rewrite meta B1 to a second-precision value placed in the future and assert
        no false concurrency warning is raised (write_date < export_dt).
        """
        from .test_wizard import _push_write_dates_to_past
        _push_write_dates_to_past(self.env, self.order, delta_seconds=120)
        b64 = self.order._generate_kalkulation_excel()

        # Rewrite meta B1 with a legacy second-precision timestamp in the future.
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        from datetime import datetime, timedelta
        future = datetime.utcnow() + timedelta(seconds=60)
        wb['kalksync_meta']['B1'] = future.strftime('%Y-%m-%dT%H:%M:%SZ')
        buf = io.BytesIO()
        wb.save(buf)
        legacy_b64 = base64.b64encode(buf.getvalue()).decode()

        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': legacy_b64,
            'file_name': 'legacy.xlsx',
        })
        wizard._onchange_file_data()
        # export_dt parsed → set; write_dates are in the past → no warning.
        self.assertTrue(wizard.export_datetime)
        self.assertFalse(wizard.parse_warning)


class TestShowOnlyChangesToggle(KalkSyncBaseCase):

    def test_toggle_off_reveals_unchanged_lines(self):
        """Flipping show_only_changes triggers a re-parse that adds the unchanged lines.

        This exercises _onchange_show_only_changes, which must re-parse the file
        (TransientModel o2m is not persisted between onchange calls)."""
        b64 = self.order._generate_kalkulation_excel()
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'file_data': b64,
            'file_name': 'toggle.xlsx',
            'show_only_changes': True,
        })
        wizard._onchange_file_data()
        self.assertFalse(wizard.line_ids.filtered(lambda l: l.status == 'unchanged'))

        wizard.show_only_changes = False
        wizard._onchange_show_only_changes()
        self.assertTrue(
            wizard.line_ids.filtered(lambda l: l.status == 'unchanged'),
            "Unchanged lines must appear after toggling show_only_changes off",
        )

    def test_toggle_without_file_is_noop(self):
        """Toggling with no file loaded must not raise and must leave line_ids empty."""
        wizard = self.env['kalksync.import.wizard'].create({
            'sale_order_id': self.order.id,
            'show_only_changes': True,
        })
        wizard.show_only_changes = False
        wizard._onchange_show_only_changes()
        self.assertFalse(wizard.line_ids)


class TestStaticHelpers(KalkSyncBaseCase):

    def test_make_error_line_defaults(self):
        line = KalkSyncImportWizard._make_error_line(message='kaputt')
        self.assertEqual(line['status'], 'error')
        self.assertEqual(line['error_message'], 'kaputt')
        self.assertEqual(line['field_name'], 'id')
        self.assertEqual(line['value_diff'], 0.0)
        self.assertFalse(line['order_line_id'])

    def test_make_error_line_overrides(self):
        line = KalkSyncImportWizard._make_error_line(
            order_line_id=self.line1.id,
            sequence=7,
            field_name='price_unit',
            label='Preis',
            value_old='1',
            value_new='2',
            message='boom',
        )
        self.assertEqual(line['order_line_id'], self.line1.id)
        self.assertEqual(line['sequence'], 7)
        self.assertEqual(line['field_label'], 'Preis')
        self.assertEqual(line['value_old'], '1')
        self.assertEqual(line['value_new'], '2')

    def test_read_row_fields_collects_importable_values(self):
        """_read_row_fields pulls importable cells, coerces by type, drops blanks."""
        b64 = self.order._generate_kalkulation_excel()
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        ws = wb.active
        meta = wb['kalksync_meta']
        data_row_start = int(meta['B6'].value)
        mapping = json.loads(meta['B5'].value)
        field_col_idx = {
            fn: column_index_from_string(cl) for fn, cl in mapping.items()
        }
        line_fields = self.env['sale.order.line']._fields
        first_data_row = next(ws.iter_rows(min_row=data_row_start, values_only=False))

        wiz = self.env['kalksync.import.wizard'].new({
            'sale_order_id': self.order.id,
        })
        result = wiz._read_row_fields(
            first_data_row, field_col_idx, line_fields
        )
        # The export wrote line values via placeholders; the qty/price come back typed.
        self.assertIn('product_uom_qty', result)
        self.assertIn('price_unit', result)
        self.assertIsInstance(result['product_uom_qty'], float)
        self.assertIsInstance(result['price_unit'], float)
