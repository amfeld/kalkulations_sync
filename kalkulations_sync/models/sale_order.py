import base64
import io
import json
import re
import zipfile
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError

from ..utils import _NON_IMPORTABLE_TYPES, _is_importable_field  # noqa: F401 (re-exported for tests)


# ---------------------------------------------------------------------------
# Module-level helpers (also importable for tests)
# ---------------------------------------------------------------------------

# Matches empty formula-cache nodes (<v/>, <v></v>, <v> </v>).
_EMPTY_VALUE_RE = re.compile(r'<v\s*/>|<v>\s*</v>')

# The SharePoint custom-properties part: harmless to drop for an exported sheet.
_SHAREPOINT_CUSTOM_PART = 'docProps/custom.xml'

# Matches an invalid W3CDTF timestamp that carries both a numeric offset and a
# trailing "Z" (e.g. "2026-06-18T10:00:00+00:00Z"). The offset is kept, the
# stray Z dropped.
_DOUBLE_TZ_RE = re.compile(r'([+-]\d{2}:\d{2})Z')


def sanitize_export_xlsx(xlsx_bytes: bytes) -> bytes:
    """Clean up openpyxl's output so Excel opens the file without a repair prompt.

    Two fixes are applied in a single repackaging pass:

    1. **Empty formula caches.** openpyxl's load(data_only=False) → save cycle
       keeps the formula string but writes an *empty* <v></v> value node (it
       discards the original cached result). Because the cache is blank, our own
       re-import reads every formula cell as None → "Formelzelle ohne
       berechneten Wert". Dropping the empty node leaves a clean <f>…</f> cell;
       with fullCalcOnLoad, Excel recalculates on open and writes real caches on
       the next save (which the import then reads).

    2. **SharePoint custom properties.** Templates that passed through SharePoint
       carry docProps/custom.xml with managed-metadata properties (ContentTypeId,
       _dlc_DocIdItemGuid, MediaServiceImageTags). openpyxl rewrites that part in
       a shape Excel rejects → "unreadable content" repair dialog (and the
       equivalent "custom XML elements no longer supported" warning in Word).
       The properties are irrelevant to an exported calculation, so the part and
       its two references (Content_Types override + root relationship) are
       removed entirely.

    3. **Invalid modified timestamp.** With the server running in UTC, openpyxl
       serialises the core "modified" property as "…+00:00Z" — an invalid
       W3CDTF value carrying both an offset and a trailing Z. Excel reports this
       during file-level validation ("repaired or discarded"). The stray Z is
       removed, leaving a valid "…+00:00".
    """
    src = io.BytesIO(xlsx_bytes)
    dst = io.BytesIO()
    with zipfile.ZipFile(src) as zin, \
            zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename
            if name == _SHAREPOINT_CUSTOM_PART:
                continue
            data = zin.read(name)
            if name.startswith('xl/worksheets/') and name.endswith('.xml'):
                data = _EMPTY_VALUE_RE.sub('', data.decode('utf-8')).encode('utf-8')
            elif name == '[Content_Types].xml':
                data = re.sub(
                    r'<Override PartName="/docProps/custom\.xml"[^>]*/>',
                    '', data.decode('utf-8'),
                ).encode('utf-8')
            elif name == '_rels/.rels':
                data = re.sub(
                    r'<Relationship[^>]*Target="docProps/custom\.xml"[^>]*/>',
                    '', data.decode('utf-8'),
                ).encode('utf-8')
            elif name == 'docProps/core.xml':
                data = _DOUBLE_TZ_RE.sub(r'\1', data.decode('utf-8')).encode('utf-8')
            zout.writestr(item, data)
    return dst.getvalue()


def adjust_formula(formula: str, row_offset: int) -> str:
    """Adjust relative row references in an Excel formula by row_offset.

    Leaves absolute row references ($A$1-style) unchanged.
    Returns the input unchanged if it is not a formula string.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def _replace(match):
        col_part = match.group(1)    # e.g. "A" or "$A"
        dollar_row = match.group(2)  # "$" or ""
        row_num = int(match.group(3))
        if dollar_row:
            return match.group(0)
        return f"{col_part}{row_num + row_offset}"

    return re.sub(r'(\$?[A-Za-z]+)(\$?)(\d+)', _replace, formula)


# Human-label → field name mapping for [label] markers in header rows.
# English and German labels are both accepted (case-insensitive).
_LABEL_TO_FIELD = {
    # English
    'quantity':           'product_uom_qty',
    'qty':                'product_uom_qty',
    'unit price':         'price_unit',
    'price per unit':     'price_unit',
    'unit cost':          'purchase_price',
    'cost per unit':      'purchase_price',
    'description':        'name',
    # German (legacy / Klarnamen)
    'menge':              'product_uom_qty',
    'preis je einheit':   'price_unit',
    'preis je me':        'price_unit',
    'kosten je einheit':  'purchase_price',
    'kosten je me':       'purchase_price',
    'bezeichnung':        'name',
}


def _resolve_header_marker(raw: str, line_fields: dict) -> str | None:
    """Resolve a [marker] value to an Odoo field name.

    Accepts technical names ([price_unit], [line.price_unit]) and
    German Klarnamen ([Preis je Einheit], [Preis je ME]) — case-insensitive.
    Returns None if the field does not exist or is not importable.
    """
    s = raw.strip().lower()
    if s.startswith('line.'):
        s = s[5:]
    resolved = _LABEL_TO_FIELD.get(s)
    if resolved and _is_importable_field(line_fields, resolved):
        return resolved
    if _is_importable_field(line_fields, s):
        return s
    return None


# Range refs must be matched before single refs (longer match wins).
_CELL_REF_RE = re.compile(
    r'(\$?[A-Za-z]+)(\$?)(\d+):(\$?[A-Za-z]+)(\$?)(\d+)'  # range
    r'|(\$?[A-Za-z]+)(\$?)(\d+)'                            # single
)


def _shift_formula_below_data(formula: str, master_row: int, offset: int) -> str:
    """Fix row references in a formula that sits below the inserted data rows.

    openpyxl's insert_rows moves cells physically but does not update formula
    strings. This repairs them:
    - Range start: unchanged (acts as anchor to data-block start)
    - Range end >= master_row: shifted by offset (SUM ranges expand)
    - Single ref >= master_row: shifted by offset
    - Absolute row refs ($) and refs < master_row: unchanged
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def _replace(m):
        if m.group(1) is not None:
            c1, d1, r1 = m.group(1), m.group(2), int(m.group(3))
            c2, d2, r2 = m.group(4), m.group(5), int(m.group(6))
            new_r2 = (r2 + offset) if (not d2 and r2 >= master_row) else r2
            return f"{c1}{d1}{r1}:{c2}{d2}{new_r2}"
        else:
            c, d, r = m.group(7), m.group(8), int(m.group(9))
            if d or r < master_row:
                return m.group(0)
            return f"{c}{r + offset}"

    return _CELL_REF_RE.sub(_replace, formula)


def _fix_below_block_formulas(ws, master_row_idx: int, offset: int) -> None:
    """After insert_rows, repair formula strings in all rows below the data block."""
    data_end_row = master_row_idx + offset
    for row in ws.iter_rows(min_row=data_end_row + 1):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = _shift_formula_below_data(
                    cell.value, master_row_idx, offset
                )


def _resolve_path(record, field_path: str):
    """Walk a dot-separated field path (max 2 levels) on an Odoo record.

    Only traverses declared Odoo fields — blocks access to private attrs
    like _cr, env, sudo(), etc.
    """
    parts = field_path.split('.')
    val = record
    for part in parts[:2]:
        if not val:
            return None
        # Only allow traversal into declared Odoo fields
        if hasattr(val, '_fields') and part not in val._fields:
            return None
        val = getattr(val, part, None)
    return val


def _to_cell_value(val, field_path: str = ''):
    """Convert an Odoo value to a native Python type suitable for a cell."""
    if val is None or val is False:
        return 0 if field_path == 'id' else None
    if hasattr(val, '_name'):  # recordset
        if len(val) == 0:
            return None
        if field_path == 'id' or field_path.endswith('.id'):
            return val.id
        return val.display_name
    if isinstance(val, bool):
        return _('Yes') if val else _('No')
    return val


def _replace_line_placeholders(cell_value, line):
    """Replace {{line.field}} placeholders in a cell value.

    Returns the native Python type when the whole cell is one placeholder,
    otherwise returns a substituted string.
    """
    if not isinstance(cell_value, str):
        return cell_value

    # Entire cell is a single placeholder → return native type
    single = re.match(r'^\s*\{\{line\.([^}]+)\}\}\s*$', cell_value)
    if single:
        field_path = single.group(1)
        return _to_cell_value(_resolve_path(line, field_path), field_path)

    # Mixed string with one or more placeholders
    def _sub(match):
        field_path = match.group(1)
        val = _resolve_path(line, field_path)
        if val is None or val is False:
            return ''
        if hasattr(val, '_name'):
            return val.display_name if len(val) == 1 else ''
        return str(val)

    return re.sub(r'\{\{line\.([^}]+)\}\}', _sub, cell_value)


# ---------------------------------------------------------------------------
# Model
# ---------------------------------------------------------------------------

class SaleOrder(models.Model):
    _inherit = 'sale.order'

    # ------------------------------------------------------------------
    # Actions called from the view
    # ------------------------------------------------------------------

    def action_export_kalkulation(self):
        """Export order lines as a filled Excel file and attach it to the chatter."""
        self.ensure_one()

        if self.state not in ('draft', 'sent'):
            raise UserError(_(
                "Export is only possible in state 'Quotation' or 'Quotation Sent'."
            ))

        template_b64 = self.company_id.amf_kalksync_template
        if not template_b64:
            raise UserError(_(
                "No calculation template configured. "
                "Please set a template under Settings → Sales."
            ))

        # Validate template contains {{line.id}}
        try:
            wb_check = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(template_b64)), data_only=False
            )
            ws_check = wb_check.active
            has_line_id = any(
                isinstance(cell.value, str) and '{{line.id}}' in cell.value
                for row in ws_check.iter_rows()
                for cell in row
            )
        except Exception as exc:
            raise UserError(_("Template could not be read: %s") % exc)

        if not has_line_id:
            raise ValidationError(_(
                "The template does not contain a {{line.id}} placeholder. "
                "Please check the template and upload it again."
            ))

        file_content = self._generate_kalkulation_excel()
        date_str = fields.Date.today().strftime('%y%m%d')
        partner_name = re.sub(r'[^\w]', '_', self.partner_id.name or '').strip('_')[:30]
        file_name = f"{date_str}_Calc_{partner_name}_{self.name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': file_content,
            'res_model': 'sale.order',
            'res_id': self.id,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })

        self.message_post(
            body=_("Calculation exported by %(user)s (%(file)s).") % {
                'user': self.env.user.name,
                'file': file_name,
            },
            attachment_ids=[attachment.id],
        )

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_import_kalkulation(self):
        """Open the import wizard pre-filled with this sale order."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Import Calculation'),
            'res_model': 'kalksync.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_sale_order_id': self.id},
        }

    # ------------------------------------------------------------------
    # Export engine
    # ------------------------------------------------------------------

    def _generate_kalkulation_excel(self):
        """Build the Excel file from the stored template. Returns base64 string."""
        self.ensure_one()

        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(self.company_id.amf_kalksync_template)),
            data_only=False,
        )
        ws = wb.active

        # --- Find master row (first row containing {{line.id}}) ---
        master_row_idx = None
        id_col_idx = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '{{line.id}}' in cell.value:
                    master_row_idx = cell.row
                    id_col_idx = cell.column
                    break
            if master_row_idx:
                break

        # --- Replace header placeholders {{object.field}} ---
        header_re = re.compile(r'\{\{object\.([^}]+)\}\}')

        def _replace_header(match):
            val = _resolve_path(self, match.group(1))
            return str(val) if val is not None and val is not False else ''

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '{{object.' in cell.value:
                    cell.value = header_re.sub(_replace_header, cell.value)

        # --- Snapshot master row (before any expansion) ---
        master_cells = []
        for cell in ws[master_row_idx]:
            master_cells.append({
                'value': cell.value,
                'number_format': cell.number_format,
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
            })

        # --- Build column mapping for importable fields ---
        # Method 1: placeholder {{line.field_name}} in master row
        # Method 2: header marker [field_name] in the row above the master row
        #           → allows formula columns (no placeholder needed in master row)
        _line_ph_re = re.compile(r'\{\{line\.([^}]+)\}\}')
        _header_marker_re = re.compile(r'\[([^\]]+)\]')  # matches [any content]
        col_mapping = {}
        line_fields = self.env['sale.order.line']._fields

        for col_idx, cd in enumerate(master_cells, start=1):
            if not isinstance(cd['value'], str):
                continue
            m = _line_ph_re.search(cd['value'])
            if m and _is_importable_field(line_fields, m.group(1)):
                col_mapping[m.group(1)] = get_column_letter(col_idx)

        # Scan all rows above the master row for [Klarname] or [field_name] markers.
        # Accepts German labels ([Preis je Einheit]) and technical names ([price_unit]).
        # Any row works — closest to master row wins (scan bottom-up).
        for scan_row in range(master_row_idx - 1, 0, -1):
            for cell in ws[scan_row]:
                if not isinstance(cell.value, str):
                    continue
                m = _header_marker_re.search(cell.value)
                if not m:
                    continue
                field = _resolve_header_marker(m.group(1), line_fields)
                if field and field not in col_mapping:
                    col_mapping[field] = get_column_letter(cell.column)

        # --- Expand rows ---
        lines = self.order_line.filtered(
            lambda l: l.display_type not in ('line_section', 'line_note')
        ).sorted(key=lambda l: (l.sequence, l.id))
        n_lines = len(lines)
        if n_lines == 0:
            raise UserError(_("No order lines available."))

        if n_lines > 1:
            ws.insert_rows(master_row_idx + 1, n_lines - 1)

        for i, line in enumerate(lines):
            row_idx = master_row_idx + i
            offset = i

            for col_idx, cd in enumerate(master_cells, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                cell.font = copy(cd['font'])
                cell.fill = copy(cd['fill'])
                cell.border = copy(cd['border'])
                cell.alignment = copy(cd['alignment'])
                cell.number_format = cd['number_format']

                value = cd['value']

                if value is None:
                    cell.value = None
                    continue

                # Adjust formula row offsets for copied rows
                if offset > 0 and isinstance(value, str) and value.startswith('='):
                    value = adjust_formula(value, offset)

                # Replace line placeholders. Sections/notes are already filtered
                # out of `lines` above, so every line here carries an importable ID.
                if isinstance(value, str) and '{{line.' in value:
                    cell.value = _replace_line_placeholders(value, line)
                else:
                    cell.value = value

        # --- Fix formula refs below data block (openpyxl insert_rows doesn't update formulas) ---
        if n_lines > 1:
            _fix_below_block_formulas(ws, master_row_idx, n_lines - 1)

        # --- Write metadata to hidden sheet ---
        if 'kalksync_meta' in wb.sheetnames:
            del wb['kalksync_meta']
        meta = wb.create_sheet('kalksync_meta')
        meta.sheet_state = 'hidden'
        meta['A1'] = 'export_datetime'
        # Microsecond precision avoids false positive concurrency warnings when
        # write_date has sub-second precision but export_dt was truncated to seconds.
        meta['B1'] = fields.Datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
        meta['A2'] = 'sale_order_id'
        meta['B2'] = self.id
        meta['A4'] = 'id_column'
        meta['B4'] = get_column_letter(id_col_idx)
        meta['A5'] = 'column_mapping'
        meta['B5'] = json.dumps(col_mapping)
        meta['A6'] = 'data_row_start'
        meta['B6'] = master_row_idx

        # Force Excel to recompute every formula on open (openpyxl drops the
        # original cached values during the load/save round-trip).
        wb.calculation.fullCalcOnLoad = True

        buf = io.BytesIO()
        wb.save(buf)
        # Repair openpyxl's output (empty formula caches + SharePoint custom
        # properties) so Excel opens the file without a corruption prompt.
        return base64.b64encode(sanitize_export_xlsx(buf.getvalue())).decode()
