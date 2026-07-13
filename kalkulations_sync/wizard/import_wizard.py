import base64
import io
import json

import openpyxl
from markupsafe import Markup
from openpyxl.utils import column_index_from_string

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools.misc import formatLang

from ..utils import _is_importable_field


_FIELD_LABELS = {
    'product_uom_qty': 'Quantity',
    'price_unit': 'Unit Price',
    'purchase_price': 'Unit Cost',
    'name': 'Description',
}

_FLOAT_TOLERANCE = 1e-6


def _coerce_field_value(f_type: str, raw):
    """Convert an Excel raw cell value to the Python type expected by Odoo write()."""
    if raw is None:
        return None
    if f_type == 'float' or f_type == 'monetary':
        if isinstance(raw, str):
            raw = raw.replace(',', '.').strip()
        return float(raw)
    if f_type == 'integer':
        if isinstance(raw, str):
            raw = raw.replace(',', '.').strip()
        return int(float(raw))
    if f_type == 'boolean':
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        return str(raw).strip().lower() in ('ja', 'yes', 'true', '1', 'wahr')
    return str(raw)  # char, text, html


def _field_label(field_name: str, f) -> str:
    # _() resolves the label against the user's language at runtime; the German
    # values live in i18n/de.po (the literals here are not auto-extracted because
    # the argument is a variable, so they are listed explicitly in the .pot).
    return _(_FIELD_LABELS.get(field_name)) if field_name in _FIELD_LABELS \
        else (f.string if f else field_name)


def _drop_internal(d):
    """Strip keys starting with '_' before passing dicts to Odoo model create()."""
    return {k: v for k, v in d.items() if not k.startswith('_')}


class KalkSyncImportWizard(models.TransientModel):
    _name = 'kalksync.import.wizard'
    _description = 'Kalkulations-Sync Import Wizard'

    sale_order_id = fields.Many2one('sale.order', required=True, readonly=True)
    file_data = fields.Binary(string='Excel file (.xlsx)')
    file_name = fields.Char()
    export_datetime = fields.Datetime(string='Export timestamp', readonly=True)
    show_only_changes = fields.Boolean(string='Show only changes', default=True)
    parse_warning = fields.Char(readonly=True)
    line_ids = fields.One2many(
        'kalksync.import.wizard.line', 'wizard_id', string='Lines',
    )
    has_errors = fields.Boolean(compute='_compute_has_errors')
    count_changed = fields.Integer(compute='_compute_counts')
    count_errors = fields.Integer(compute='_compute_counts')
    count_new = fields.Integer(compute='_compute_counts')
    count_missing = fields.Integer(compute='_compute_counts')

    @api.depends('line_ids.status')
    def _compute_has_errors(self):
        for rec in self:
            rec.has_errors = any(l.status == 'error' for l in rec.line_ids)

    @api.depends('line_ids.status')
    def _compute_counts(self):
        for rec in self:
            statuses = rec.line_ids.mapped('status')
            rec.count_changed = statuses.count('changed')
            rec.count_errors = statuses.count('error')
            rec.count_new = statuses.count('new')
            rec.count_missing = statuses.count('missing')

    @api.onchange('file_data', 'file_name')
    def _onchange_file_data(self):
        if not self.file_data:
            self.line_ids = [(5,)]
            self.export_datetime = False
            self.parse_warning = False
            return

        try:
            line_vals, export_dt, warnings = self._parse_excel()
        except UserError as exc:
            self.line_ids = [(5,)]
            self.export_datetime = False
            return {'warning': {'title': _('Import error'), 'message': str(exc)}}

        self.line_ids = [(5,)] + [
            (0, 0, _drop_internal(v)) for v in self._filter_lines(line_vals)
        ]
        self.export_datetime = export_dt
        self.parse_warning = '\n'.join(warnings) if warnings else False

    @api.onchange('show_only_changes')
    def _onchange_show_only_changes(self):
        if not self.file_data:
            return
        # Re-parse is necessary because line_ids on a TransientModel are not
        # persisted between onchange calls — the DB rows don't exist yet.
        try:
            line_vals, _dt, _w = self._parse_excel()
        except UserError:
            return
        self.line_ids = [(5,)] + [
            (0, 0, _drop_internal(v)) for v in self._filter_lines(line_vals)
        ]

    def _filter_lines(self, line_vals):
        if self.show_only_changes:
            return [v for v in line_vals if v.get('status') != 'unchanged']
        return line_vals

    # ------------------------------------------------------------------
    # Parse
    # ------------------------------------------------------------------

    def _parse_excel(self):
        """Return (wizard_line_vals, export_datetime, warning_strings)."""
        raw = base64.b64decode(self.file_data)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise UserError(_("File could not be opened: %s") % exc)

        # Second load without data_only so we can detect formula cells
        # whose cached results are missing (file not saved after calculation).
        try:
            wb_formulas = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
            ws_formulas = wb_formulas.active
        except Exception:
            ws_formulas = None

        if 'kalksync_meta' not in wb.sheetnames:
            raise UserError(_(
                "This file was not exported with Kalkulations-Sync "
                "(missing 'kalksync_meta' sheet)."
            ))

        meta = wb['kalksync_meta']

        try:
            export_datetime_str = meta['B1'].value or ''
            sale_order_id = int(meta['B2'].value)
            id_col_letter = str(meta['B4'].value)
            col_mapping_json = meta['B5'].value or '{}'
            data_row_start = int(meta['B6'].value)
        except Exception as exc:
            raise UserError(_("Metadata could not be read: %s") % exc)

        if sale_order_id != self.sale_order_id.id:
            raise UserError(_(
                "This file was exported for quotation ID %(file_id)d, "
                "not for the current quotation (ID %(curr_id)d)."
            ) % {'file_id': sale_order_id, 'curr_id': self.sale_order_id.id})

        # Parse export timestamp — support both second-precision (legacy files)
        # and microsecond-precision (current format).
        try:
            from datetime import datetime as _dt
            dt_str = export_datetime_str.replace('Z', '').replace('T', ' ')
            fmt = '%Y-%m-%d %H:%M:%S.%f' if '.' in dt_str else '%Y-%m-%d %H:%M:%S'
            export_dt = _dt.strptime(dt_str, fmt)
        except Exception:
            export_dt = False

        # Column mapping: field_name → column letter
        try:
            col_mapping = json.loads(col_mapping_json)
        except Exception:
            col_mapping = {}

        id_col_idx = column_index_from_string(id_col_letter)
        field_col_idx = {
            fn: column_index_from_string(cl)
            for fn, cl in col_mapping.items()
        }

        line_fields = self.env['sale.order.line']._fields
        available_fields = [fn for fn in field_col_idx if _is_importable_field(line_fields, fn)]

        order = self.sale_order_id
        so_lines_by_id = {line.id: line for line in order.order_line}
        excel_line_ids = set()
        wizard_lines = []
        warnings = []

        ws = wb.active

        for row in ws.iter_rows(min_row=data_row_start, values_only=False):
            # Skip entirely empty rows
            if all(cell.value is None for cell in row):
                continue

            id_cell = row[id_col_idx - 1] if id_col_idx <= len(row) else None
            raw_id = id_cell.value if id_cell else None

            # Row with no ID (sum rows, notes, spacers) → skip silently
            if raw_id is None:
                continue

            # 'N' or 'n' → neue Position anlegen
            if str(raw_id).strip().lower() == 'n':
                new_fields = self._read_row_fields(row, field_col_idx, line_fields)
                display_name = new_fields.get('name') or _('New line (from calculation)')
                wizard_lines.append({
                    'order_line_id': False,
                    'sequence': 0,
                    'field_name': '_new_line',
                    'field_label': _('New line'),
                    'value_old': '',
                    'value_new': display_name,
                    'value_diff': 0.0,
                    'status': 'new',
                    'error_message': '',
                    '_new_line_fields': new_fields,
                })
                continue

            # Non-integer ID → manipulation
            try:
                line_id = int(raw_id)
            except (ValueError, TypeError):
                wizard_lines.append(self._make_error_line(
                    label=_('Line ID'),
                    value_new=str(raw_id),
                    message=_("The line ID was modified. This row cannot be imported."),
                ))
                continue

            # Duplicate ID → user copied this row → treat as new position
            if line_id in excel_line_ids:
                new_fields = self._read_row_fields(row, field_col_idx, line_fields)
                display_name = new_fields.get('name') or _('New line (copy)')
                wizard_lines.append({
                    'order_line_id': False,
                    'sequence': 0,
                    'field_name': '_new_line',
                    'field_label': _('New line (copy)'),
                    'value_old': '',
                    'value_new': display_name,
                    'value_diff': 0.0,
                    'status': 'new',
                    'error_message': '',
                    '_new_line_fields': new_fields,
                })
                continue

            excel_line_ids.add(line_id)

            # ID not found in SO
            if line_id not in so_lines_by_id:
                wizard_lines.append(self._make_error_line(
                    label=_('Line ID'),
                    value_new=str(line_id),
                    message=_("Line ID %d is not present in the quotation.") % line_id,
                ))
                continue

            so_line = so_lines_by_id[line_id]

            # Section lines: mark as ignored (only name editable, handled separately)
            if so_line.display_type == 'line_section':
                wizard_lines.append({
                    'order_line_id': so_line.id,
                    'sequence': so_line.sequence,
                    'field_name': 'display_type',
                    'field_label': _('Section'),
                    'value_old': so_line.name or '',
                    'value_new': so_line.name or '',
                    'value_diff': 0.0,
                    'status': 'ignored',
                    'error_message': '',
                })
                continue

            # Concurrency check
            if export_dt and so_line.write_date and so_line.write_date > export_dt:
                warnings.append(
                    _("Line '%s' was modified in Odoo after the export.")
                    % (so_line.product_id.display_name or str(so_line.id))
                )

            actual_row_idx = row[0].row

            # Compare each importable field
            for field_name in available_fields:
                if field_name not in field_col_idx:
                    continue
                col_idx = field_col_idx[field_name]
                if col_idx > len(row):
                    continue

                f = line_fields.get(field_name)
                f_type = f.type if f else 'float'
                label = _field_label(field_name, f)
                odoo_val = getattr(so_line, field_name, None)
                excel_raw = row[col_idx - 1].value

                if excel_raw is None:
                    # Check whether this is a formula cell whose result was
                    # never cached (file not opened+saved in Excel after editing).
                    if ws_formulas is not None:
                        formula_val = ws_formulas.cell(
                            row=actual_row_idx, column=col_idx
                        ).value
                        if isinstance(formula_val, str) and formula_val.startswith('='):
                            odoo_display = self._fmt_num(float(odoo_val or 0)) \
                                if f_type in ('float', 'integer', 'monetary') \
                                else str(odoo_val or '')
                            wizard_lines.append(self._make_error_line(
                                order_line_id=so_line.id,
                                sequence=so_line.sequence,
                                field_name=field_name,
                                label=label,
                                value_old=odoo_display,
                                value_new=formula_val,
                                message=_(
                                    "Formula cell without a calculated value ('%s'). "
                                    "Please open the file in Microsoft Excel, "
                                    "save it and import again."
                                ) % formula_val,
                            ))
                            continue

                if f_type in ('float', 'integer', 'monetary'):
                    if excel_raw is None:
                        # Blank cell → no change intended (explicit 0 is returned as 0, not None)
                        continue
                    raw_norm = str(excel_raw).replace(',', '.').strip() \
                        if isinstance(excel_raw, str) else excel_raw
                    try:
                        excel_coerced = float(raw_norm)
                    except (ValueError, TypeError):
                        wizard_lines.append(self._make_error_line(
                            order_line_id=so_line.id,
                            sequence=so_line.sequence,
                            field_name=field_name,
                            label=label,
                            value_old=self._fmt_num(float(odoo_val or 0)),
                            value_new=str(excel_raw),
                            message=_(
                                "Invalid value '%s' for field '%s'."
                            ) % (excel_raw, label),
                        ))
                        continue
                    odoo_float = float(odoo_val or 0)
                    diff = excel_coerced - odoo_float
                    status = 'unchanged' if abs(diff) < _FLOAT_TOLERANCE else 'changed'
                    wizard_lines.append({
                        'order_line_id': so_line.id,
                        'sequence': so_line.sequence,
                        'field_name': field_name,
                        'field_label': label,
                        'value_old': self._fmt_num(odoo_float),
                        'value_new': self._fmt_num(excel_coerced),
                        'value_diff': diff,
                        'status': status,
                        'error_message': '',
                        # Raw float preserved so action_confirm avoids the
                        # 6-decimal fmt/parse round-trip (_drop_internal strips this).
                        '_raw_value': excel_coerced,
                    })
                elif f_type == 'boolean':
                    excel_bool = _coerce_field_value('boolean', excel_raw)
                    odoo_bool = bool(odoo_val)
                    status = 'unchanged' if excel_bool == odoo_bool else 'changed'
                    wizard_lines.append({
                        'order_line_id': so_line.id,
                        'sequence': so_line.sequence,
                        'field_name': field_name,
                        'field_label': label,
                        'value_old': _('Yes') if odoo_bool else _('No'),
                        'value_new': _('Yes') if excel_bool else _('No'),
                        'value_diff': 0.0,
                        'status': status,
                        'error_message': '',
                        '_raw_value': excel_bool,
                    })
                elif f_type == 'selection':
                    # Selection keys differ between instances/versions, so an
                    # unvalidated write would raise ValueError at confirm time
                    # and roll back ALL changes. Accept the technical key or
                    # the (translated) label, map to the key, and turn invalid
                    # values into a per-row error instead.
                    if excel_raw is None or str(excel_raw).strip() == '':
                        # Blank cell → no change intended (like blank numbers)
                        continue
                    sel_pairs = f._description_selection(self.env)
                    label_by_key = dict(sel_pairs)
                    excel_key = self._selection_key(f, excel_raw)
                    odoo_key = odoo_val or ''
                    if excel_key is None:
                        wizard_lines.append(self._make_error_line(
                            order_line_id=so_line.id,
                            sequence=so_line.sequence,
                            field_name=field_name,
                            label=label,
                            value_old=label_by_key.get(odoo_key, ''),
                            value_new=str(excel_raw),
                            message=_(
                                "Invalid value '%(val)s' for field '%(field)s'. "
                                "Allowed values: %(allowed)s"
                            ) % {
                                'val': excel_raw,
                                'field': label,
                                'allowed': ', '.join(
                                    str(lbl) for _k, lbl in sel_pairs
                                ),
                            },
                        ))
                        continue
                    status = 'unchanged' if excel_key == odoo_key else 'changed'
                    wizard_lines.append({
                        'order_line_id': so_line.id,
                        'sequence': so_line.sequence,
                        'field_name': field_name,
                        'field_label': label,
                        'value_old': label_by_key.get(odoo_key, ''),
                        'value_new': label_by_key.get(excel_key, excel_key),
                        'value_diff': 0.0,
                        'status': status,
                        'error_message': '',
                        '_raw_value': excel_key,
                    })
                else:  # char, text, html
                    excel_str = str(excel_raw) if excel_raw is not None else ''
                    odoo_str = str(odoo_val) if odoo_val else ''
                    status = 'unchanged' if excel_str == odoo_str else 'changed'
                    wizard_lines.append({
                        'order_line_id': so_line.id,
                        'sequence': so_line.sequence,
                        'field_name': field_name,
                        'field_label': label,
                        'value_old': odoo_str,
                        'value_new': excel_str,
                        'value_diff': 0.0,
                        'status': status,
                        'error_message': '',
                        '_raw_value': excel_str,
                    })

        # Lines in Odoo but missing in Excel
        for line_id, so_line in so_lines_by_id.items():
            if line_id not in excel_line_ids \
                    and so_line.display_type not in ('line_section', 'line_note'):
                wizard_lines.append({
                    'order_line_id': so_line.id,
                    'sequence': so_line.sequence,
                    'field_name': False,
                    'field_label': _('Missing line'),
                    'value_old': so_line.product_id.display_name or str(so_line.id),
                    'value_new': '',
                    'value_diff': 0.0,
                    'status': 'missing',
                    'error_message': '',
                })
                warnings.append(
                    _("Line '%s' is not present in the Excel file.")
                    % (so_line.product_id.display_name or str(so_line.id))
                )

        return wizard_lines, export_dt, warnings

    # ------------------------------------------------------------------
    # Confirm
    # ------------------------------------------------------------------

    def action_confirm(self):
        """Write changed values to sale.order.line and create new lines."""
        self.ensure_one()

        # Re-parse the file at confirm time. The line_ids one2many is populated
        # only via @api.onchange (display-only), so it is NOT saved to the DB
        # and cannot be read back here. Parsing again is the reliable path.
        try:
            line_vals, _dt, _warnings = self._parse_excel()
        except UserError:
            raise

        if any(v.get('status') == 'error' for v in line_vals):
            raise UserError(_(
                "There are still rows with errors. Please correct the Excel "
                "file and upload it again."
            ))

        so_lines_by_id = {line.id: line for line in self.sale_order_id.order_line}
        changes_by_line = {}
        for v in line_vals:
            if v.get('status') != 'changed':
                continue
            line_id = v.get('order_line_id')
            field_name = v.get('field_name')
            if not line_id or not field_name:
                continue
            so_line = so_lines_by_id.get(int(line_id))
            if not so_line:
                continue
            f = self.env['sale.order.line']._fields.get(field_name)
            f_type = f.type if f else 'float'
            # Use raw value from parse to avoid fmt/parse round-trip precision loss.
            if '_raw_value' in v:
                val = v['_raw_value']
            else:
                try:
                    val = _coerce_field_value(f_type, v['value_new'])
                except (ValueError, TypeError):
                    val = None
            if val is not None:
                changes_by_line.setdefault(so_line, {})[field_name] = val

        # Each line needs potentially different values, so individual writes are
        # unavoidable here — grouping by identical vals would yield no benefit
        # for typical pricing/quantity changes across distinct positions.
        # kalksync_import unterdrückt die per-Feld-Chatter-Notizen aus sale_wpr;
        # die Änderungen werden weiter unten in eine Nachricht zusammengefasst.
        for ol, vals in changes_by_line.items():
            ol.with_context(kalksync_import=True).write(vals)

        n_updated = len(changes_by_line)

        # Create new lines (rows marked with 'N' in the ID column)
        new_count = 0
        new_vals_list = [v for v in line_vals if v.get('status') == 'new']
        if new_vals_list:
            default_product = self.sale_order_id.company_id.amf_kalksync_default_product_id
            max_seq = max(
                (l.sequence for l in self.sale_order_id.order_line), default=0
            )
            for v in new_vals_list:
                fields_vals = dict(v.get('_new_line_fields') or {})
                if 'product_id' not in fields_vals:
                    if not default_product:
                        raise UserError(_(
                            "New line '%(name)s' cannot be created: "
                            "no product assigned and no default product configured.\n"
                            "Tip: set a default product under Settings → Sales → "
                            "Kalkulations-Sync."
                        ) % {'name': fields_vals.get('name') or _('Unknown line')})
                    fields_vals['product_id'] = default_product.id
                    if default_product.uom_id and 'product_uom_id' not in fields_vals:
                        fields_vals['product_uom_id'] = default_product.uom_id.id

                max_seq += 10
                name = fields_vals.pop('name', _('New line (from calculation)'))
                self.env['sale.order.line'].create({
                    'order_id': self.sale_order_id.id,
                    'name': name,
                    'sequence': max_seq,
                    **fields_vals,
                })
                new_count += 1

        now = fields.Datetime.now()

        # Attach original file to chatter, prefixed with U<JJMMTT>_ so the upload
        # is easy to spot in the attachments list.
        attachment_ids = []
        if self.file_data and self.file_name:
            att = self.env['ir.attachment'].create({
                'name': 'U%s_%s' % (now.strftime('%y%m%d'), self.file_name),
                'type': 'binary',
                'datas': self.file_data,
                'res_model': 'sale.order',
                'res_id': self.sale_order_id.id,
                'mimetype': (
                    'application/vnd.openxmlformats-officedocument'
                    '.spreadsheetml.sheet'
                ),
            })
            attachment_ids = [att.id]

        new_part = (_(", %(n)d newly created") % {'n': new_count}) if new_count else ''
        body = _(
            "Kalkulations-Sync import: %(n)d line(s) updated"
            "%(new_part)s on %(date)s by %(user)s."
        ) % {
            'n': n_updated,
            'new_part': new_part,
            'date': now.strftime('%d.%m.%Y %H:%M'),
            'user': self.env.user.name,
        }

        # Fold the individual field changes into the same message instead of
        # letting each become its own chatter note (see sale_wpr write override).
        # message_post escapes plain str bodies, so the HTML must be a Markup.
        html_body = Markup('<p>%s</p>') % body
        detail = self._format_change_details(line_vals, so_lines_by_id)
        if detail:
            html_body += detail

        self.sale_order_id.message_post(
            body=html_body, attachment_ids=attachment_ids,
        )

        return {'type': 'ir.actions.act_window_close'}

    def _format_change_details(self, line_vals, so_lines_by_id):
        """Render the changed fields grouped per order line as one HTML block.

        Returns a Markup so message_post keeps the HTML intact; all
        interpolated values are escaped by markupsafe.
        """
        currency = self.sale_order_id.currency_id
        line_fields = self.env['sale.order.line']._fields

        def _fmt(f, s):
            if f and f.type in ('float', 'monetary'):
                try:
                    return formatLang(self.env, float(s), currency_obj=currency)
                except (ValueError, TypeError):
                    return s
            return s

        grouped = {}  # so_line -> [(field label, old, new)]
        for v in line_vals:
            if v.get('status') != 'changed':
                continue
            so_line = so_lines_by_id.get(v.get('order_line_id'))
            if not so_line:
                continue
            f = line_fields.get(v.get('field_name'))
            grouped.setdefault(so_line, []).append((
                v.get('field_label') or v.get('field_name'),
                _fmt(f, v.get('value_old') or '0'),
                _fmt(f, v.get('value_new') or '0'),
            ))

        blocks = []
        for so_line, changes in grouped.items():
            # Multi-line descriptions: the first line is enough as heading.
            title = (so_line.name or str(so_line.id)).splitlines()[0]
            items = Markup('').join(
                Markup(
                    '<li>%s: <span style="color:#888888;">%s</span>'
                    ' → <strong>%s</strong></li>'
                ) % (label, old, new)
                for label, old, new in changes
            )
            blocks.append(
                Markup(
                    '<p style="margin:8px 0 0 0;"><strong>%s</strong></p>'
                    '<ul style="margin:0 0 4px 0;">%s</ul>'
                ) % (title, items)
            )
        return Markup('').join(blocks)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _read_row_fields(self, row, field_col_idx, line_fields):
        """Collect importable field values from an Excel row into a dict."""
        result = {}
        for fn, col_idx in field_col_idx.items():
            if not _is_importable_field(line_fields, fn):
                continue
            if col_idx > len(row):
                continue
            raw = row[col_idx - 1].value
            if raw is None:
                continue
            f = line_fields.get(fn)
            if f is not None and f.type == 'selection':
                # Invalid keys would raise ValueError at create() and roll back
                # the whole import — drop them here (blank = field not set).
                val = self._selection_key(f, raw)
            else:
                try:
                    val = _coerce_field_value(f.type if f else 'float', raw)
                except (ValueError, TypeError):
                    val = None
            if val is not None:
                result[fn] = val
        return result

    def _selection_key(self, f, raw):
        """Map an Excel cell value to a valid selection key.

        Accepts the technical key or the translated label, case-insensitive.
        Returns None if the value matches neither.
        """
        s = str(raw).strip()
        sel_pairs = f._description_selection(self.env)
        if s in {key for key, _lbl in sel_pairs}:
            return s
        lower = s.lower()
        for key, lbl in sel_pairs:
            if str(lbl).strip().lower() == lower:
                return key
        return None

    @staticmethod
    def _fmt_num(val):
        if val == 0.0:
            return '0'
        s = f'{val:.6f}'.rstrip('0').rstrip('.')
        return s

    @staticmethod
    def _make_error_line(**kw):
        return {
            'order_line_id': kw.get('order_line_id', False),
            'sequence': kw.get('sequence', 0),
            'field_name': kw.get('field_name', 'id'),
            'field_label': kw.get('label', _('Error')),
            'value_old': kw.get('value_old', ''),
            'value_new': kw.get('value_new', ''),
            'value_diff': 0.0,
            'status': 'error',
            'error_message': kw.get('message', ''),
        }


class KalkSyncImportWizardLine(models.TransientModel):
    _name = 'kalksync.import.wizard.line'
    _description = 'Kalkulations-Sync Import Wizard – Line'
    _order = 'sequence, field_name'

    wizard_id = fields.Many2one('kalksync.import.wizard', required=True, ondelete='cascade')
    order_line_id = fields.Many2one('sale.order.line', string='Order line')
    sequence = fields.Integer(string='Pos.')
    field_name = fields.Char(string='Field name')
    field_label = fields.Char(string='Field')
    value_old = fields.Char(string='Odoo value', readonly=True)
    value_new = fields.Char(string='Excel value', readonly=True)
    value_diff = fields.Float(string='Difference', digits=(16, 6), readonly=True)
    status = fields.Selection([
        ('error', 'Error'),
        ('changed', 'Changed'),
        ('unchanged', 'Unchanged'),
        ('ignored', 'Ignored'),
        ('missing', 'Missing'),
        ('new', 'New'),
    ], string='Status')
    error_message = fields.Char(string='Error message', readonly=True)
